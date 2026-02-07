"""
Excel spreadsheet (.xlsx) text extraction using openpyxl.

Converts spreadsheet data to semantic text format for better RAG search.
Each row becomes a descriptive line using column headers.
"""

import io
import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def extract_xlsx_text(source: bytes | str) -> str:
    """
    Extract text content from an Excel spreadsheet (.xlsx).

    Converts tabular data to readable text format:
    - Uses first row as column headers
    - Each data row becomes "Row N: Header1=Value1, Header2=Value2, ..."
    - Handles multiple worksheets

    Args:
        source: Raw XLSX file content (bytes) or path to XLSX file (str)

    Returns:
        Extracted text with semantic row descriptions.
        Returns empty string if extraction fails.
    """
    try:
        # Support both bytes and file paths
        file_input = io.BytesIO(source) if isinstance(source, bytes) else source
        wb = load_workbook(file_input, read_only=True, data_only=True)
        text_parts: list[str] = []

        sheet_count = len(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = _extract_sheet_text(sheet, sheet_name)
            if sheet_text:
                text_parts.append(sheet_text)

        wb.close()

        full_text = "\n\n".join(text_parts)
        logger.info("Extracted %d characters from XLSX (%d sheets)", len(full_text), sheet_count)
        return full_text

    except Exception as e:
        logger.error("Failed to extract XLSX: %s", e)
        return ""


def _extract_sheet_text(sheet, sheet_name: str) -> str:
    """
    Extract text from a single worksheet.

    Converts rows to semantic descriptions using column headers.
    """
    lines: list[str] = [f"Sheet: {sheet_name}"]

    # Read all rows
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    # Filter out completely empty rows
    rows = [row for row in rows if any(cell is not None for cell in row)]
    if not rows:
        return ""

    # Use first row as headers
    headers = []
    for i, cell in enumerate(rows[0]):
        if cell is not None:
            headers.append(str(cell).strip())
        else:
            headers.append(f"Column{i+1}")

    # Process data rows
    for row_idx, row in enumerate(rows[1:], start=1):
        pairs = []
        for col_idx, cell in enumerate(row):
            if cell is not None:
                cell_str = str(cell).strip()
                if cell_str:
                    header = headers[col_idx] if col_idx < len(headers) else f"Column{col_idx+1}"
                    pairs.append(f"{header}={cell_str}")

        if pairs:
            lines.append(f"Row {row_idx}: {', '.join(pairs)}")

    # Only return if we have data rows (not just header)
    if len(lines) > 1:
        return "\n".join(lines)
    return ""


def extract_xlsx_raw(source: bytes | str) -> str:
    """
    Extract raw text from Excel without semantic formatting.

    Alternative extraction that just concatenates all cell values.
    Useful when semantic format isn't needed.
    """
    try:
        file_input = io.BytesIO(source) if isinstance(source, bytes) else source
        wb = load_workbook(file_input, read_only=True, data_only=True)
        text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"=== {sheet_name} ===")

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip(" |"):
                    text_parts.append(row_text)

        wb.close()
        return "\n".join(text_parts)

    except Exception as e:
        logger.error("Failed to extract XLSX (raw): %s", e)
        return ""
